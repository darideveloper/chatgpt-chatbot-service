import os
import openpyxl
from business_data import models


def read_excel(file_path: os.path, sheet_name: str) -> list:
    """ Extract data from xlsx file and return a list or rows

    Args:
        file_path (os.path): File path with data
        sheet_name (str): Sheet name with data
        
    Returns:
        list: List of rows with data
    """
    
    # Open excel file
    wb = openpyxl.load_workbook(file_path)
    sheet = wb[sheet_name]
    
    # Read data from file
    rows = sheet.max_row
    columns = sheet.max_column
    data = []
    more_data = True
    for row in range(1, rows + 1):

        row_data = []
        for column in range(1, columns + 1):
            cell_data = sheet.cell(row, column).value
            
            if column == 1 and cell_data is None:
                more_data = False
                break
            
            row_data.append(cell_data)

        if not more_data:
            break

        data.append(row_data)
        
    return data


def refaccionaria_x_xlsx(file_path: os.path):
    """ Extract data from xlsx file and save to database

    Args:
        file_path (os.path): File path with data
    """
    
    data = read_excel(file_path, "database")
        
    # Delete old products
    models.RefaccionariaX.objects.all().delete()
        
    # Save data to database
    for row in data[1:]:
        models.RefaccionariaX.objects.create(
            id=row[0],
            nombre=row[1].strip(),
            descripcion=row[2].strip(),
            fabricante=row[3].strip(),
            numero_de_pieza=row[4].strip(),
            categoria=row[5].strip(),
            precio=row[6],
            cantidad_en_stock=row[7],
            ubicacion=row[8].strip(),
            estante=row[9].strip(),
            modelo=row[10].strip(),
            ano=row[11].strip(),
        )
        
    print("Done!")


def refaccionaria_y_xlsx(file_path: os.path):
    """ Extract data from xlsx file and save to database

    Args:
        file_path (os.path): File path with data
    """
    
    data = read_excel(file_path, "database")
        
    # Delete old products
    models.RefaccionariaY.objects.all().delete()
        
    # Save data to database
    for row in data[1:]:
        models.RefaccionariaY.objects.create(
            id=row[0],
            nombre=row[1].strip(),
            descripcion=row[2].strip(),
            fabricante=row[3].strip(),
            numero_de_pieza=row[4].strip(),
            categoria=row[5].strip(),
            precio=row[6],
            cantidad_en_stock=row[7],
            ubicacion=row[8].strip(),
            estante=row[9].strip(),
            modelo=row[10].strip(),
            ano=row[11].strip(),
        )
        
    print("Done!")
    

def refaccionaria_gonzalez_xlsx(file_path: os.path):
    """ Extract data from xlsx file and save to database

    Args:
        file_path (os.path): File path with data
    """
    
    data = read_excel(file_path, "Hoja1")
        
    # Save data to database
    for row in data[1:]:
        models.RefaccionariaGonzalez.objects.create(
            name=row[1].strip(),
            sku=str(row[2]).strip(),
            brand=row[3].strip(),
            origin=row[4].strip(),
            compatibility=row[5].strip(),
            price=row[6],
            category=row[7].strip()
        )
        
    print("Done!")
    
    
# Relation between business name and file extension
FILES_RELATION = {
    "refaccionaria x": {
        "xlsx": refaccionaria_x_xlsx,
    },
    "refaccionaria y": {
        "xlsx": refaccionaria_y_xlsx,
    },
    "refaccionaria gonzalez": {
        "xlsx": refaccionaria_gonzalez_xlsx,
    },
}