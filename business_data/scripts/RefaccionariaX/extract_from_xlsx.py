import os
import sys
import django
import openpyxl

# paths
CURRENT_FOLDER = os.path.dirname(__file__)
SCRIPTS_FOLDER = os.path.dirname(CURRENT_FOLDER)
APP_FOLDER = os.path.dirname(SCRIPTS_FOLDER)
PROJECT_FOLDER = os.path.dirname(APP_FOLDER)
DATA_FILES_FOLDER = os.path.join(APP_FOLDER, "data_files")

sys.path.append(APP_FOLDER)
sys.path.append(PROJECT_FOLDER)

# Setup django settings
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'chatgpt_chatbot_service.settings')
django.setup()

from business_data.models import RefaccionariaX

# Open excel file
excel_file = os.path.join(DATA_FILES_FOLDER, "RefaccionariaX.xlsx")
wb = openpyxl.load_workbook(excel_file)
sheet = wb["database"]

# Read data from file
rows = sheet.max_row
columns = sheet.max_column
data = []
for row in range(1, rows + 1):

    row_data = []
    for column in range(1, columns + 1):
        cell_data = sheet.cell(row, column).value
        row_data.append(cell_data)

    data.append(row_data)
    
# Delete old products
RefaccionariaX.objects.all().delete()
    
# Save data to database
for row in data[1:]:
    RefaccionariaX.objects.create(
        id=row[0],
        nombre=row[1].strip(),
        descripcion=row[2].strip(),
        fabricante=row[3].strip(),
        numero_de_pieza=row[4].strip(),
        categoria=row[5].strip(),
        precio=row[6],
        cantidad_en_stock=row[7],
        ubicacion=row[8].strip(),
        estante=row[9].strip(),
        modelo=row[10].strip(),
        ano=row[11].strip(),
    )
    
print("Done!")